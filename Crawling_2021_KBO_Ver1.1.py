import sys
import os
import numpy as np
import openpyxl
import pandas as pd
import time
from selenium import webdriver
from selenium.webdriver.support.ui import Select
from openpyxl import load_workbook


t_date=input()
t_year=int(float(t_date[0:4]))
t_month=int(float(t_date[4:6]))
t_day=int(float(t_date[6:8]))
if getattr(sys, 'frozen', False):
    t_path = os.path.dirname(sys.executable)
elif __file__:
    t_path = os.path.dirname(__file__)
openpyxl.Workbook().save(t_path+'\%s.xlsx' %t_date)
book=load_workbook(t_path+'\%s.xlsx' %t_date)
next_h_row=0
next_p_row=0
next_e_col=0

# selenium에서 사용할 웹 드라이버 절대 경로 정보
chromedriver = os.path.join(t_path,'chromedriver.exe')
# selenum의 webdriver에 앞서 설치한 chromedirver를 연동한다.
driver = webdriver.Chrome(chromedriver)
# driver로 특정 페이지를 크롤링한다.
driver.get('https://www.koreabaseball.com/Schedule/GameCenter/Main.aspx')
time.sleep(1)

open_calender=driver.find_element_by_xpath('//*[@id="contents"]/div[2]/ul/li[2]/img')
open_calender.click()

year_select=Select(driver.find_element_by_xpath('//*[@id="ui-datepicker-div"]/div/div/select[2]'))
year_select.select_by_value(str(t_year))

month_select=Select(driver.find_element_by_xpath('//*[@id="ui-datepicker-div"]/div/div/select[1]'))
month_select.select_by_value(str(t_month-1))

day_pick=driver.find_element_by_link_text(str(t_day))
day_pick.click()
time.sleep(1)

for i in range(1,6):
    if driver.find_elements_by_xpath('//*[@id="contents"]/div[3]/div/div[1]/ul/li[%s][@game_sc="3"]' %i):
        driver.find_element_by_xpath('//*[@id="contents"]/div[3]/div/div[1]/ul/li[%s][@game_sc="3"]' %i).click()
        AwayTeam=driver.find_element_by_xpath('//*[@id="contents"]/div[3]/div/div[1]/ul/li[%s][@game_sc="3"]' %i).get_attribute('away_nm')
        HomeTeam=driver.find_element_by_xpath('//*[@id="contents"]/div[3]/div/div[1]/ul/li[%s][@game_sc="3"]' %i).get_attribute('home_nm')
        driver.find_element_by_xpath('//*[@id="tabDepth2"]/li[2]').click()
        time.sleep(1)

        #기타
        etc_li=driver.find_element_by_xpath('//*[@id="tblEtc"]').get_attribute('outerHTML')
        etc=np.concatenate(pd.read_html(etc_li))
        df_etc=pd.DataFrame(etc)



        #타자
        A_hitter1_li=driver.find_element_by_xpath('//*[@id="tblAwayHitter1"]').get_attribute('outerHTML')
        A_hitter1=np.concatenate(pd.read_html(A_hitter1_li))
        df_A_hitter1=pd.DataFrame(A_hitter1)

        A_hitter2_li=driver.find_element_by_xpath('//*[@id="tblAwayHitter2"]/table').get_attribute('outerHTML')
        A_hitter2=np.concatenate(pd.read_html(A_hitter2_li))
        df_A_hitter2=pd.DataFrame(A_hitter2)

        A_hitter3_li=driver.find_element_by_xpath('//*[@id="tblAwayHitter3"]').get_attribute('outerHTML')
        A_hitter3=np.concatenate(pd.read_html(A_hitter3_li))
        df_A_hitter3=pd.DataFrame(A_hitter3)

        A_hitter=pd.concat([df_A_hitter1,df_A_hitter2], axis=1)


        B_hitter1_li=driver.find_element_by_xpath('//*[@id="tblHomeHitter1"]').get_attribute('outerHTML')
        B_hitter1=np.concatenate(pd.read_html(B_hitter1_li))
        df_B_hitter1=pd.DataFrame(B_hitter1)

        B_hitter2_li=driver.find_element_by_xpath('//*[@id="tblHomeHitter2"]/table').get_attribute('outerHTML')
        B_hitter2=np.concatenate(pd.read_html(B_hitter2_li))
        df_B_hitter2=pd.DataFrame(B_hitter2)

        B_hitter3_li=driver.find_element_by_xpath('//*[@id="tblHomeHitter3"]').get_attribute('outerHTML')
        B_hitter3=np.concatenate(pd.read_html(B_hitter3_li))
        df_B_hitter3=pd.DataFrame(B_hitter3)

        B_hitter=pd.concat([df_B_hitter1,df_B_hitter2], axis=1)


        AB_hitter=pd.concat([A_hitter[:-1],B_hitter[:-1]])
        AB_hitter_end=pd.concat([df_A_hitter3[:-1],df_B_hitter3[:-1]])



        #투수
        A_pitcher_li=driver.find_element_by_xpath('//*[@id="tblAwayPitcher"]').get_attribute('outerHTML')
        A_pitcher=np.concatenate(pd.read_html(A_pitcher_li))
        df_A_pitcher=pd.DataFrame(A_pitcher)

        B_pitcher_li=driver.find_element_by_xpath('//*[@id="tblHomePitcher"]').get_attribute('outerHTML')
        B_pitcher=np.concatenate(pd.read_html(B_pitcher_li))
        df_B_pitcher=pd.DataFrame(B_pitcher)

        AB_pitcher=pd.concat([df_A_pitcher[:-1],df_B_pitcher[:-1]])

        h_row=len(AB_hitter)
        p_row=len(AB_pitcher)
        e_col=len(df_etc.columns)



        #저장
        with pd.ExcelWriter(t_path+'\%s.xlsx' %t_date, engine='openpyxl', mode='a') as writer:
            writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
            df_etc.to_excel(writer,sheet_name='기타기록',startcol=next_e_col, header=False, index=False)
            AB_hitter.to_excel(writer,sheet_name='타자기록',startrow=next_h_row, header=False, index=False)
            AB_hitter_end.to_excel(writer,sheet_name='타자기록',startrow=next_h_row, startcol=15, header=False, index=False)
            AB_pitcher.to_excel(writer,sheet_name='투수기록',startrow=next_p_row, header=False, index=False)
            AB_pitcher.to_excel(writer,sheet_name=str(i)+HomeTeam+AwayTeam,startrow=0, header=False, index=False)
            AB_hitter.to_excel(writer,sheet_name=str(i)+HomeTeam+AwayTeam,startrow=max(p_row,len(df_etc))+1, header=False, index=False)
            AB_hitter_end.to_excel(writer,sheet_name=str(i)+HomeTeam+AwayTeam,startrow=max(p_row,len(df_etc))+1, startcol=15, header=False, index=False)
            df_etc.to_excel(writer,sheet_name=str(i)+HomeTeam+AwayTeam,startcol=len(AB_pitcher.columns)+1, header=False, index=False)

            next_h_row += h_row
            next_p_row += p_row
            next_e_col += e_col
            
            
    else:
        continue

if driver.find_element_by_link_text('Next'):
    driver.find_element_by_link_text('Next').click()
    for i in range(6,11):
        if driver.find_elements_by_xpath('//*[@id="contents"]/div[3]/div/div[1]/ul/li[%s][@game_sc="3"]' %i):
            driver.find_element_by_xpath('//*[@id="contents"]/div[3]/div/div[1]/ul/li[%s][@game_sc="3"]' %i).click()
            AwayTeam=driver.find_element_by_xpath('//*[@id="contents"]/div[3]/div/div[1]/ul/li[%s][@game_sc="3"]' %i).get_attribute('away_nm')
            HomeTeam=driver.find_element_by_xpath('//*[@id="contents"]/div[3]/div/div[1]/ul/li[%s][@game_sc="3"]' %i).get_attribute('home_nm')
            driver.find_element_by_xpath('//*[@id="tabDepth2"]/li[2]').click()
            time.sleep(1)

            #기타
            etc_li=driver.find_element_by_xpath('//*[@id="tblEtc"]').get_attribute('outerHTML')
            etc=np.concatenate(pd.read_html(etc_li))
            df_etc=pd.DataFrame(etc)



            #타자
            A_hitter1_li=driver.find_element_by_xpath('//*[@id="tblAwayHitter1"]').get_attribute('outerHTML')
            A_hitter1=np.concatenate(pd.read_html(A_hitter1_li))
            df_A_hitter1=pd.DataFrame(A_hitter1)

            A_hitter2_li=driver.find_element_by_xpath('//*[@id="tblAwayHitter2"]/table').get_attribute('outerHTML')
            A_hitter2=np.concatenate(pd.read_html(A_hitter2_li))
            df_A_hitter2=pd.DataFrame(A_hitter2)

            A_hitter3_li=driver.find_element_by_xpath('//*[@id="tblAwayHitter3"]').get_attribute('outerHTML')
            A_hitter3=np.concatenate(pd.read_html(A_hitter3_li))
            df_A_hitter3=pd.DataFrame(A_hitter3)

            A_hitter=pd.concat([df_A_hitter1,df_A_hitter2], axis=1)


            B_hitter1_li=driver.find_element_by_xpath('//*[@id="tblHomeHitter1"]').get_attribute('outerHTML')
            B_hitter1=np.concatenate(pd.read_html(B_hitter1_li))
            df_B_hitter1=pd.DataFrame(B_hitter1)

            B_hitter2_li=driver.find_element_by_xpath('//*[@id="tblHomeHitter2"]/table').get_attribute('outerHTML')
            B_hitter2=np.concatenate(pd.read_html(B_hitter2_li))
            df_B_hitter2=pd.DataFrame(B_hitter2)

            B_hitter3_li=driver.find_element_by_xpath('//*[@id="tblHomeHitter3"]').get_attribute('outerHTML')
            B_hitter3=np.concatenate(pd.read_html(B_hitter3_li))
            df_B_hitter3=pd.DataFrame(B_hitter3)

            B_hitter=pd.concat([df_B_hitter1,df_B_hitter2], axis=1)


            AB_hitter=pd.concat([A_hitter[:-1],B_hitter[:-1]])
            AB_hitter_end=pd.concat([df_A_hitter3[:-1],df_B_hitter3[:-1]])



            #투수
            A_pitcher_li=driver.find_element_by_xpath('//*[@id="tblAwayPitcher"]').get_attribute('outerHTML')
            A_pitcher=np.concatenate(pd.read_html(A_pitcher_li))
            df_A_pitcher=pd.DataFrame(A_pitcher)

            B_pitcher_li=driver.find_element_by_xpath('//*[@id="tblHomePitcher"]').get_attribute('outerHTML')
            B_pitcher=np.concatenate(pd.read_html(B_pitcher_li))
            df_B_pitcher=pd.DataFrame(B_pitcher)

            AB_pitcher=pd.concat([df_A_pitcher[:-1],df_B_pitcher[:-1]])

            h_row=len(AB_hitter)
            p_row=len(AB_pitcher)
            e_col=len(df_etc.columns)



            #저장
            with pd.ExcelWriter(t_path+'\%s.xlsx' %t_date, engine='openpyxl', mode='a') as writer:
                writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                df_etc.to_excel(writer,sheet_name='기타기록',startcol=next_e_col, header=False, index=False)
                AB_hitter.to_excel(writer,sheet_name='타자기록',startrow=next_h_row, header=False, index=False)
                AB_hitter_end.to_excel(writer,sheet_name='타자기록',startrow=next_h_row, startcol=15, header=False, index=False)
                AB_pitcher.to_excel(writer,sheet_name='투수기록',startrow=next_p_row, header=False, index=False)
                AB_pitcher.to_excel(writer,sheet_name=str(i)+HomeTeam+AwayTeam,startrow=0, header=False, index=False)
                AB_hitter.to_excel(writer,sheet_name=str(i)+HomeTeam+AwayTeam,startrow=max(p_row,len(df_etc))+1, header=False, index=False)
                AB_hitter_end.to_excel(writer,sheet_name=str(i)+HomeTeam+AwayTeam,startrow=max(p_row,len(df_etc))+1, startcol=15, header=False, index=False)
                df_etc.to_excel(writer,sheet_name=str(i)+HomeTeam+AwayTeam,startcol=len(AB_pitcher.columns)+1, header=False, index=False)

                next_h_row += h_row
                next_p_row += p_row
                next_e_col += e_col
                
                
        else:
            continue

